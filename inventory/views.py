import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from .models import Items, Stock, Suppliers, Categories, Orders, OrderItems, Attachments, Warehouses, Units
from .forms import ItemForm, ProductFilterForm, StockForm, SupplierForm, OrderForm, OrderItemFormSet, OrderItemForm
from django.http import HttpResponseRedirect, HttpResponse
from django.utils.text import format_lazy
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import render
from django.db.models import Q, Count, F, IntegerField, OuterRef, Subquery
from itertools import product
from django.db import models
import json
from django.http import JsonResponse
from django.db.models.functions import Coalesce
from django.db.models import DecimalField, Value, FloatField
from openpyxl import Workbook
from urllib.parse import urlencode
from django.template.loader import render_to_string
from openpyxl.styles import Font
from django.forms import inlineformset_factory
from django.db import transaction
from datetime import datetime
from django.forms import modelformset_factory
from django.contrib import messages
from django.contrib.contenttypes.models import ContentType
import re
from pathlib import Path
from django.core.paginator import Paginator
from django.utils.timezone import now



def home(request):
    return render(request, 'inventory/home.html')

def item_list(request):
    # Pobranie liczby elementów na stronę z domyślną wartością
    items_per_page = request.GET.get('showing', '10').strip()
    try:
        items_per_page = int(items_per_page) if items_per_page.isdigit() else 10
    except ValueError:
        items_per_page = 10  # Ustaw domyślną wartość w przypadku błędu

    # Filtry
    name_query = request.GET.get('name', '').strip()
    ean_query = request.GET.get('ean', '').strip()
    price_query = request.GET.get('price', '').strip()
    category_query = request.GET.get('category', '').strip()
    supplier_query = request.GET.get('supplier', '').strip()
    minquantity_query = request.GET.get('minquantity', '').strip()
    unit_query = request.GET.get('unit', '').strip()
    isactive_query = request.GET.get('isactive', '').strip()


    items = Items.objects.select_related('itm_catid', 'itm_supid', 'itm_uniid').all()

    # Filtry
    if name_query:
        items = items.filter(itm_name__icontains=name_query)
    if ean_query:
        items = items.filter(itm_ean__icontains=ean_query)
    if price_query:
        try:
            items = items.filter(itm_price=float(price_query))
        except ValueError:
            pass
    if category_query:
        items = items.filter(itm_catid__cat_name__icontains=category_query)
    if supplier_query:
        items = items.filter(itm_supid__sup_name__icontains=supplier_query)
    if minquantity_query:
        try:
            items = items.filter(itm_minquantity=int(minquantity_query))
        except ValueError:
            pass
    if unit_query:
        items = items.filter(itm_uniid__uni_name__icontains=unit_query)
    if isactive_query == "1":  # Aktywny
        items = items.filter(itm_isactive=True)
    elif isactive_query == "0":  # Nieaktywny
        items = items.filter(itm_isactive=False)

    # Paginacja
    page_number = request.GET.get('page', 1)
    paginator = Paginator(items, items_per_page)
    page_obj = paginator.get_page(page_number)

    # dynamiczne ładowanie tabeli
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        table_html = render_to_string('inventory/item_list_table.html', {'items': page_obj})
        pagination_html = render_to_string('inventory/item_list_pagination.html', {'items': page_obj})
        return JsonResponse({'html': table_html, 'pagination': pagination_html})

    return render(request, 'inventory/item_list.html', {
        'items': page_obj,
        'items_per_page': items_per_page,
        'name_query': name_query,
        'ean_query': ean_query,
        'price_query': price_query,
        'category_query': category_query,
        'supplier_query': supplier_query,
        'minquantity_query': minquantity_query,
        'unit_query': unit_query,
        'isactive_query': isactive_query,
    })

def add_item(request):
    if request.method == 'POST':
        form = ItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)  # Nie zapisuj jeszcze rekordu
            item.save(user=request.user)  # Przekaż aktualnie zalogowanego użytkownika
            return redirect('item_list')
    else:
        form = ItemForm()

    # Pobranie wszystkich kategorii i dostawców
    categories = Categories.objects.all()
    suppliers = Suppliers.objects.all()
    units = Units.objects.all()

    return render(request, 'inventory/add_item.html', {
        'form': form,
        'categories': categories,
        'suppliers': suppliers,
        'units': units,
    })


def edit_item(request, item_id):
    item = get_object_or_404(Items, pk=item_id)

    if request.method == "POST":
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('item_list')
    else:
        form = ItemForm(instance=item)

    return render(request, 'inventory/edit_item.html', {
        'form': form,
        'item': item,
    })
def delete_item(request, item_id):
    # Znajdź produkt na podstawie jego ID
    item = get_object_or_404(Items, itm_id=item_id)

    # Usuń produkt
    item.delete()

    # Po usunięciu, przekieruj użytkownika z powrotem do listy produktów
    return redirect('item_list')

from django.shortcuts import render
from .models import Items

def ajax_filter_items(request):
    items = Items.objects.all()

    name = request.GET.get('name')
    ean = request.GET.get('ean')
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')

    # Filtrowanie
    if name:
        items = items.filter(itm_name__icontains=name)
    if ean:
        items = items.filter(itm_ean__icontains=ean)
    if min_price:
        items = items.filter(itm_price__gte=min_price)
    if max_price:
        items = items.filter(itm_price__lte=max_price)

    return render(request, 'inventory/item_list_table.html', {'items': items})

from django.core.paginator import Paginator
from django.shortcuts import render
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.db.models import Q
from .models import Stock

def stock_list(request):
    # Pobranie liczby elementów na stronę (domyślnie 10)
    stocks_per_page = request.GET.get('showing', '10').strip()
    try:
        stocks_per_page = int(stocks_per_page) if stocks_per_page.isdigit() else 10
    except ValueError:
        stocks_per_page = 10  # Domyślna wartość w przypadku błędu

    # Filtry
    product_query = request.GET.get('product', '').strip()
    warehouse_query = request.GET.get('warehouse', '').strip()
    quantity_query = request.GET.get('quantity', '').strip()

    # Pobranie danych z bazy
    stocks = Stock.objects.select_related('stk_itmid', 'stk_whsid').all()

    # Filtrowanie
    if product_query:
        stocks = stocks.filter(stk_itmid__itm_name__icontains=product_query)
    if warehouse_query:
        stocks = stocks.filter(stk_whsid__whs_name__icontains=warehouse_query)
    if quantity_query:
        try:
            stocks = stocks.filter(stk_qty=int(quantity_query))
        except ValueError:
            pass

    # Paginacja
    page_number = request.GET.get('page', 1)
    paginator = Paginator(stocks, stocks_per_page)
    page_obj = paginator.get_page(page_number)

    # AJAX – dynamiczne ładowanie tabeli i paginacji
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        table_html = render_to_string('inventory/stock_list_table.html', {'stocks': page_obj})
        pagination_html = render_to_string('inventory/stock_list_pagination.html', {'stocks': page_obj})
        return JsonResponse({'html': table_html, 'pagination': pagination_html})

    return render(request, 'inventory/stock_list.html', {
        'stocks': page_obj,
        'stocks_per_page': stocks_per_page,
        'product_query': product_query,
        'warehouse_query': warehouse_query,
        'quantity_query': quantity_query,
    })

def add_stock(request):
    if request.method == 'POST':
        form = StockForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock_list')
    else:
        form = StockForm()
    return render(request, 'inventory/add_stock.html', {'form': form})


def edit_stock(request, stock_id):
    stock = get_object_or_404(Stock, pk=stock_id)


    if request.method == 'POST':
        form = StockForm(request.POST, instance=stock)
        if form.is_valid():
            form.save()
            return redirect('stock_list')
    else:
        form = StockForm(instance=stock)

    return render(request, 'inventory/edit_stock.html', {'form': form})


def delete_stock(request, stock_id):
    stock = get_object_or_404(Stock, pk=stock_id)
    stock.delete()
    return redirect('stock_list')


def supplier_list(request):
    showing = int(request.GET.get('showing', 10))
    page_number = request.GET.get('page', 1)

    # Pobieranie filtrów
    name_query = request.GET.get('name', '').strip()
    nip_query = request.GET.get('nip', '').strip()
    email_query = request.GET.get('email', '').strip()
    phone_query = request.GET.get('phone', '').strip()
    address_query = request.GET.get('address', '').strip()

    # Filtrowanie dostawców
    suppliers = Suppliers.objects.all()
    if name_query:
        suppliers = suppliers.filter(sup_name__icontains=name_query)
    if nip_query:
        suppliers = suppliers.filter(sup_taxid__icontains=nip_query)
    if email_query:
        suppliers = suppliers.filter(sup_email__icontains=email_query)
    if phone_query:
        suppliers = suppliers.filter(sup_phone__icontains=phone_query)
    if address_query:
        suppliers = suppliers.filter(sup_address__icontains=address_query)

    # Paginacja
    paginator = Paginator(suppliers, showing)
    try:
        page_obj = paginator.get_page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.get_page(1)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # Zapytanie AJAX
        table_html = render_to_string('inventory/supplier_list_table.html', {'suppliers': page_obj})
        pagination_html = render_to_string('inventory/supplier_list_pagination.html', {'suppliers': page_obj, 'showing': showing})
        return JsonResponse({'html': table_html, 'pagination': pagination_html})

    return render(request, 'inventory/supplier_list.html', {
        'suppliers': page_obj,
        'showing': showing,
        'name_query': name_query,
        'nip_query': nip_query,
        'email_query': email_query,
        'phone_query': phone_query,
        'address_query': address_query,
    })


def add_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'inventory/add_supplier.html', {'form': form})


def edit_supplier(request, supplier_id):
    supplier = get_object_or_404(Suppliers, pk=supplier_id)

    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)

    return render(request, 'inventory/edit_supplier.html', {'form': form})


def delete_supplier(request, supplier_id):
    # Pobierz dostawcę lub zwróć 404, jeśli nie istnieje
    supplier = get_object_or_404(Suppliers, pk=supplier_id)
    # Usuń dostawcę
    supplier.delete()
    # Przekieruj do listy dostawców
    return redirect('supplier_list')

def order_list(request):
    showing = int(request.GET.get('showing', 10))
    page_number = request.GET.get('page', 1)

    # Filtry
    number_query = request.GET.get('ord_number', '').strip()
    date_query = request.GET.get('ord_date', '').strip()
    warehouse_query = request.GET.get('ord_whsid', '').strip()
    status_query = request.GET.get('ord_status', '').strip()
    supplier_query = request.GET.get('ord_supid', '').strip()

    # Filtrowanie
    orders = Orders.objects.all()
    if number_query:
        orders = orders.filter(ord_number__icontains=number_query)
    if date_query:
        orders = orders.filter(ord_date__icontains=date_query)
    if warehouse_query:
        orders = orders.filter(ord_whsid__whs_name__icontains=warehouse_query)
    if status_query:
        orders = orders.filter(ord_statusid__sta_name__icontains=status_query)
    if supplier_query:
        orders = orders.filter(ord_supid__sup_name__icontains=supplier_query)

    # Paginacja
    paginator = Paginator(orders, showing)
    try:
        page_obj = paginator.get_page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.get_page(1)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # Zapytanie AJAX
        table_html = render_to_string('inventory/order_list_table.html', {'orders': page_obj})
        pagination_html = render_to_string('inventory/order_list_pagination.html', {
            'orders': page_obj,
            'showing': showing,
        })
        return JsonResponse({'html': table_html, 'pagination': pagination_html})

    return render(request, 'inventory/order_list.html', {
        'orders': page_obj,
        'showing': showing,
        'number_query': number_query,
        'date_query': date_query,
        'warehouse_query': warehouse_query,
        'status_query': status_query,
        'supplier_query': supplier_query,
    })


def add_order(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            with transaction.atomic():
                current_time = now()
                year = current_time.strftime("%y")
                month = current_time.strftime("%m")

                last_order = Orders.objects.filter(
                    ord_date__year=current_time.year,
                    ord_date__month=current_time.month
                ).order_by('-ord_id').first()

                next_number = 1
                if last_order and last_order.ord_number:
                    match = re.search(r'(\d+)$', last_order.ord_number)
                    if match:
                        next_number = int(match.group(1)) + 1

                order_number = f"ORD/{year}/{month}/{next_number:03d}"

                order = Orders.objects.create(
                    ord_number=order_number,
                    ord_date=data.get("order_date", current_time),
                    ord_statusid_id=data.get("order_status"),
                    ord_whsid_id=data.get("warehouse"),
                    ord_supid_id=data.get("supplier"),
                )

                for item in data.get("items", []):
                    OrderItems.objects.create(
                        oit_ordid=order,
                        oit_itmid_id=item["item"],
                        oit_quantity=item["quantity"],
                        oit_price=item["price"]
                    )

            return JsonResponse({"success": True, "order_number": order_number})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    else:
        form = OrderForm()
        items = Items.objects.all()
        return render(request, 'inventory/add_order.html', {
            'form': form,
            'items': items,
        })

def edit_order(request, order_id):
    order = get_object_or_404(Orders, pk=order_id)
    OrderItemFormset = modelformset_factory(OrderItems, form=OrderItemForm, extra=0)

    if request.method == "POST":
        form = OrderForm(request.POST, instance=order)
        formset = OrderItemFormset(request.POST, queryset=OrderItems.objects.filter(oit_ordid=order))

        if form.is_valid() and formset.is_valid():
            # Zapisz zamówienie
            order = form.save()

            # Zapisz pozycje zamówienia
            instances = formset.save(commit=False)
            for instance in instances:
                instance.oit_ordid = order  # Przypisz zamówienie do pozycji
                instance.save()

            # Usuń pozycje oznaczone do usunięcia
            for instance in formset.deleted_objects:
                instance.delete()

            return JsonResponse({"success": True})
        else:
            # Jeśli formularz jest nieprawidłowy, zwróć błędy
            errors = {
                "form_errors": form.errors,
                "formset_errors": formset.errors,
            }
            return JsonResponse({"success": False, "errors": errors})

    else:
        # GET request: wyświetl formularz do edycji
        form = OrderForm(instance=order)
        formset = OrderItemFormset(queryset=OrderItems.objects.filter(oit_ordid=order))
        all_products = Items.objects.all()

        return render(request, "inventory/edit_order.html", {
            "form": form,
            "formset": formset,
            "all_products": all_products,
            "order_id": order_id,  # Przekazanie order_id do szablonu
        })
def delete_order(request, order_id):
    # Pobierz dostawcę lub zwróć 404, jeśli nie istnieje
    order = get_object_or_404(Orders, pk=order_id)
    # Usuń dostawcę
    order.delete()
    # Przekieruj do listy dostawców
    return redirect('order_list')  # Użyj nazwy ścieżki zdefiniowanej w urls.py

def dashboard(request):
    # Liczba zamówień według statusów
    order_statuses = Orders.objects.values('ord_statusid__sta_name').annotate(count=Count('ord_statusid'))

    # Pobierz ID wybranego magazynu z GET i przekształć na liczbę (jeśli istnieje)
    selected_warehouse_id = request.GET.get('warehouse')
    try:
        selected_warehouse_id = int(selected_warehouse_id) if selected_warehouse_id else None
    except ValueError:
        selected_warehouse_id = None  # Jeśli wartość nie jest liczbą, ustaw na None

    # Pobierz wszystkie magazyny (do dropdown)
    warehouses = Warehouses.objects.all()

    # Pobieranie tylko aktywne produkty
    items = Items.objects.filter(itm_isactive=True)

    # Filtrowanie dla dashboard_data
    dashboard_data = []
    for warehouse in warehouses:
        if selected_warehouse_id and warehouse.whs_id != selected_warehouse_id:
            continue  # Jeśli wybrano magazyn, pomiń inne

        for item in items:
            stock = Stock.objects.filter(stk_itmid=item, stk_whsid=warehouse).first()
            dashboard_data.append({
                'item_name': item.itm_name,
                'warehouse_name': warehouse.whs_name,
                'actual_qty': stock.stk_qty if stock else 0,
                'min_qty': item.itm_minquantity
            })

    # Kontekst do szablonu
    context = {
        'dashboard_data': dashboard_data,
        'warehouses': warehouses,  # Wszystkie magazyny dla dropdown
        'selected_warehouse_id': selected_warehouse_id,  # Wybrany magazyn jako liczba
        'order_statuses': order_statuses,
    }
    return render(request, 'inventory/dashboard.html', context)
def export_items_to_excel(request):
    # Pobierz filtry z request.GET
    name = request.GET.get('name', '').strip()
    ean = request.GET.get('ean', '').strip()
    price = request.GET.get('price', '').strip()
    category = request.GET.get('category', '').strip()
    supplier = request.GET.get('supplier', '').strip()
    minquantity = request.GET.get('minquantity', '').strip()
    unit = request.GET.get('unit', '').strip()
    isactive = request.GET.get('isactive', '').strip()

    # Filtruj produkty na podstawie podanych parametrów
    items = Items.objects.select_related('itm_catid', 'itm_supid', 'itm_uniid')

    if name:
        items = items.filter(itm_name__icontains=name)
    if ean:
        items = items.filter(itm_ean__icontains=ean)
    if price:
        try:
            items = items.filter(itm_price=float(price))
        except ValueError:
            pass  # Ignoruj błędy konwersji, jeśli wartość jest nieprawidłowa
    if category:
        items = items.filter(itm_catid__cat_name__icontains=category)
    if supplier:
        items = items.filter(itm_supid__sup_name__icontains=supplier)
    if minquantity:
        try:
            items = items.filter(itm_minquantity=int(minquantity))
        except ValueError:
            pass
    if unit:
        items = items.filter(itm_uniid__uni_name__icontains=unit)
    # Filtr Aktywności
    if isactive == "1":  # Jeśli wybrano "Tak"
        items = items.filter(itm_isactive=1)
    elif isactive == "0":  # Jeśli wybrano "Nie"
        items = items.filter(itm_isactive=0)

    # Utwórz nowy skoroszyt
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produkty"

    # Dodaj nagłówki do pierwszego wiersza
    headers = ['Nazwa', 'Kod EAN', 'Cena', 'Kategoria', 'Dostawca', 'Ilość Minimalna', 'Jednostka Miary', 'Aktywny']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)  # Nagłówki pogrubione

    # Dodaj dane produktów
    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=item.itm_name)
        ws.cell(row=row_num, column=2, value=item.itm_ean)
        ws.cell(row=row_num, column=3, value=item.itm_price)
        ws.cell(row=row_num, column=4, value=item.itm_catid.cat_name if item.itm_catid else '-')
        ws.cell(row=row_num, column=5, value=item.itm_supid.sup_name if item.itm_supid else '-')
        ws.cell(row=row_num, column=6, value=item.itm_minquantity)
        ws.cell(row=row_num, column=7, value=item.itm_uniid.uni_name if item.itm_uniid else '-')
        ws.cell(row=row_num, column=8, value='Tak' if item.itm_isactive else 'Nie')

    # Ustaw odpowiedź HTTP z plikiem Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Produkty.xlsx'

    # Zapisz skoroszyt do odpowiedzi
    wb.save(response)
    return response
def export_suppliers_to_excel(request):
    # Pobierz filtry z request.GET
    name = request.GET.get('name', '')
    nip = request.GET.get('nip', '')
    email = request.GET.get('email', '')
    phone = request.GET.get('phone', '')
    address = request.GET.get('address', '')

    # Filtruj dostawców na podstawie podanych parametrów
    suppliers = Suppliers.objects.filter(
        Q(sup_name__icontains=name) &
        Q(sup_taxid__icontains=nip) &
        Q(sup_email__icontains=email) &
        Q(sup_phone__icontains=phone) &
        Q(sup_address__icontains=address)
    )

    # Utwórz nowy skoroszyt
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dostawcy"

    # Dodaj nagłówki do pierwszego wiersza
    headers = ['Nazwa Dostawcy', 'NIP', 'Email', 'Telefon', 'Adres']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)  # Nagłówki pogrubione

    # Dodaj dane dostawców
    for row_num, supplier in enumerate(suppliers, 2):
        ws.cell(row=row_num, column=1, value=supplier.sup_name)
        ws.cell(row=row_num, column=2, value=supplier.sup_taxid)
        ws.cell(row=row_num, column=3, value=supplier.sup_email)
        ws.cell(row=row_num, column=4, value=supplier.sup_phone)
        ws.cell(row=row_num, column=5, value=supplier.sup_address)

    # Ustaw odpowiedź HTTP z plikiem Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Dostawcy.xlsx'

    # Zapisz skoroszyt do odpowiedzi
    wb.save(response)
    return response

def export_stocks_to_excel(request):
    product_query = request.GET.get('product', '').strip()
    warehouse_query = request.GET.get('warehouse', '').strip()
    quantity_query = request.GET.get('quantity', '').strip()

    stocks = Stock.objects.all()
    if product_query:
        stocks = stocks.filter(stk_itmid__itm_name__icontains=product_query)
    if warehouse_query:
        stocks = stocks.filter(stk_whsid__whs_name__icontains=warehouse_query)
    if quantity_query:
        stocks = stocks.filter(stk_qty__icontains=quantity_query)

    # Tworzenie pliku Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stany Magazynowe"

    # Nagłówki
    ws.append(["Produkt", "Magazyn", "Ilość"])

    # Dane
    for stock in stocks:
        ws.append([
            stock.stk_itmid.itm_name or '',
            stock.stk_whsid.whs_name or '',
            stock.stk_qty or 0
        ])

    # Zapis do pliku
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="stany_magazynowe.xlsx"'
    wb.save(response)
    return response


def export_orders_to_excel(request):
    # Pobierz parametry GET

    number_query = request.GET.get('ord_number', '').strip()
    date_query = request.GET.get('ord_date', '').strip()
    warehouse_query = request.GET.get('ord_whsid', '').strip()
    status_query = request.GET.get('ord_status', '').strip()
    supplier_query = request.GET.get('ord_supid', '').strip()

    orders = Orders.objects.select_related('ord_statusid', 'ord_whsid', 'ord_supid').all()
    if number_query:
        orders = orders.filter(ord_number__icontains=number_query)
    if date_query:
        orders = orders.filter(ord_date__icontains=date_query)
    if warehouse_query:
        orders = orders.filter(ord_whsid__whs_name__icontains=warehouse_query)
    if status_query:
        orders = orders.filter(ord_statusid__sta_name__icontains=status_query)
    if supplier_query:
        orders = orders.filter(ord_supid__sup_name__icontains=supplier_query)

    # Tworzenie arkusza Excela
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zamówienia"

    # Nagłówki
    headers = ['Numer Zamówienia', 'Data', 'Magazyn', 'Status', 'Dostawca', 'Kwota']
    ws.append(headers)

    # Dane
    for order in orders:
        ws.append([
            order.ord_number,
            order.ord_date.strftime('%Y-%m-%d %H:%M'),
            order.ord_whsid.whs_name,
            order.ord_statusid.sta_name,
            order.ord_supid.sup_name if order.ord_supid else '-',
            float(order.ord_total)
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=zamowienia.xlsx'
    wb.save(response)
    return response


